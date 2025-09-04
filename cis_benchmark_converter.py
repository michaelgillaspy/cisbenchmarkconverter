#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
File: cis_benchmark_converter.py
Author: Maxime Beauchamp
LinkedIn: https://www.linkedin.com/in/maxbeauchamp/
Created: 2024-11-06
Last Update: 2025-28-04

Description:
    This script extracts recommendations from CIS Benchmark PDF documents and exports
    them to CSV, Excel, or JSON format. The extraction starts at a configurable page number 
    (to skip table of contents or disclaimers). Logging level and other parameters 
    are configurable via command-line options.

Usage:
    python cis_benchmark_converter.py \
        -i path/to/input_file.pdf \
        -o path/to/output_file \
        -f [csv|excel|json] \
        --start_page 10 \
        --log_level INFO

Dependencies:
    - pdfplumber : for text extraction from PDF files.
    - openpyxl   : for creating and handling Excel files.
    - tqdm       : for displaying a progress bar.
    - logging    : part of the standard Python library.
    - pathlib    : part of the standard Python library.
    - json       : part of the standard Python library.

Installation:
    pip install pdfplumber openpyxl tqdm

License:
    This script is provided under the MIT License.
    Please respect the copyright of the CIS Benchmarks
    documents when using and sharing this script.
"""

import argparse
import csv
import json
import re
import logging
from pathlib import Path
from typing import Tuple, List, Dict

import pdfplumber
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# Disable warnings from pdfminer and pdfplumber
# to avoid cluttering the output with unnecessary messages.
logging.getLogger("pdfminer").setLevel(logging.ERROR)
logging.getLogger("pdfplumber").setLevel(logging.ERROR)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# -----------------------------------------------------------------------------------
# Global Constants and Regular Expressions
# -----------------------------------------------------------------------------------

# Matches recommendation titles (e.g., "1.1.1 (L1) Title of Recommendation")
TITLE_PATTERN: re.Pattern = re.compile(r'^(\d+\.\d+(?:\.\d+)*)\s*(\(L\d+\))?\s*(.*)')

# Matches page number strings (e.g., "Page 123")
PAGE_NUMBER_PATTERN: re.Pattern = re.compile(r'\bPage\s+\d+\b', re.IGNORECASE)

# List of section headers to extract
SECTIONS_WITHOUT_CIS: List[str] = [
    'Profile Applicability:',
    'Description:',
    'Rationale:',
    'Impact:',
    'Audit:',
    'Remediation:',
    'Default Value:',
    'References:',
    'Additional Information:'
]

# -----------------------------------------------------------------------------------
# Utility Functions
# -----------------------------------------------------------------------------------

def remove_page_numbers(text: str) -> str:
    """
    Remove mentions of page numbers (e.g. "Page 123") from the provided text.
    """
    return PAGE_NUMBER_PATTERN.sub('', text)

def generate_unique_filename(base_name: str, extension: str) -> str:
    """
    Generate a unique filename by appending a numeric suffix if the file already exists.
    Uses pathlib for robust path handling.
    """
    file_path = Path(f"{base_name}.{extension}")
    counter = 1
    while file_path.exists():
        file_path = Path(f"{base_name}({counter}).{extension}")
        counter += 1
    return str(file_path)

# -----------------------------------------------------------------------------------
# PDF Extraction Functions
# -----------------------------------------------------------------------------------

def extract_title_and_version(input_file: Path) -> Tuple[str, str]:
    """
    Extract the document title and version from the first page of the PDF.
    
    Returns:
        (title, version) as strings. Version may be empty if no version line is found.
    """
    try:
        with pdfplumber.open(str(input_file)) as pdf:
            first_page_text = pdf.pages[0].extract_text().splitlines()
    except Exception as e:
        logging.error(f"Error opening PDF for title extraction: {e}")
        raise

    title_lines: List[str] = []
    version: str = ""
    for line in first_page_text:
        # Example: "v1.2 - 2024"
        if line.lower().startswith("v") and "-" in line:
            version = line.strip()
            break
        else:
            title_lines.append(line.strip())

    title = " ".join(title_lines) if title_lines else "CIS Benchmark Document"
    return title, version

def read_pdf(input_file: Path, start_page: int = 10) -> str:
    """
    Reads text from the PDF file starting at 'start_page'. 
    Uses tqdm to display a progress bar for the pages processed.

    Raises:
        ValueError: if 'start_page' is out of range.
        Exception:  if reading fails for another reason.

    Returns:
        A single string containing the concatenated text of the pages read.
    """
    logging.info(f"Reading PDF from page {start_page} onwards...")

    if start_page < 1:
        raise ValueError("start_page must be >= 1.")

    try:
        with pdfplumber.open(str(input_file)) as pdf:
            total_pages = len(pdf.pages)
            if start_page > total_pages:
                raise ValueError(f"Start page {start_page} exceeds total page count ({total_pages}).")

            # Extract text from each page starting at 'start_page'
            text_pages = []
            for page in tqdm(pdf.pages[start_page - 1:], 
                             desc="Extracting pages", 
                             unit="page", 
                             total=(total_pages - start_page + 1)):
                page_text = page.extract_text() or ""
                text_pages.append(page_text)

    except Exception as e:
        logging.error(f"Failed to read PDF: {e}")
        raise

    # Filter out any empty strings and join with newlines
    return "\n".join(filter(None, text_pages))

def find_profile_applicability(lines: List[str], start_index: int, max_depth: int = 10) -> bool:
    """
    Checks if 'Profile Applicability:' appears within 'max_depth' lines 
    after 'start_index', indicating a valid recommendation start.
    
    Returns:
        True if found, otherwise False.
    """
    for i in range(start_index + 1, min(start_index + max_depth, len(lines))):
        line: str = lines[i].strip()
        if line.startswith("Profile Applicability:"):
            return True
        if TITLE_PATTERN.match(line) or any(line.startswith(sec) for sec in SECTIONS_WITHOUT_CIS):
            return False
    return False

def extract_section(lines: List[str], start_index: int, section_name: str) -> Tuple[str, int]:
    """
    Extract the content of a section until a new section header, a new recommendation title,
    or a mention of "CIS Controls" is encountered.
    
    Returns:
        (content, next_index) 
        content     : the extracted text
        next_index  : the position in 'lines' after extraction
    """
    content: List[str] = []
    current_index: int = start_index + 1

    while current_index < len(lines):
        line: str = lines[current_index].strip()
        line = remove_page_numbers(line)

        # End of this section if:
        #   - We reach another known section header
        #   - We detect a new recommendation title
        #   - We see "CIS Controls"
        if any(line.startswith(sec) for sec in SECTIONS_WITHOUT_CIS) \
           or TITLE_PATTERN.match(line) \
           or line.lower().startswith("cis controls"):
            break

        content.append(line)
        current_index += 1

    # Special handling for References section to preserve URLs
    if section_name == "References:":
        # Join lines intelligently to preserve URLs
        result = []
        for line in content:
            if result and (line.startswith("http") or (result[-1].endswith("/") and not line.startswith("http"))):
                # If current line starts with http or previous ends with /, concatenate without space
                result[-1] += line
            elif result and result[-1].endswith("-"):
                # Handle hyphenated URLs that break across lines
                result[-1] = result[-1][:-1] + line
            else:
                result.append(line)
        return ' '.join(result).strip(), current_index
    else:
        return ' '.join(content).strip(), current_index

def extract_recommendations(full_text: str) -> List[Dict[str, str]]:
    """
    Parse the concatenated PDF text to extract recommendations.

    Returns:
        A list of dictionaries, each representing a recommendation with keys like
        "Number", "Level", "Title", and the extracted sections (Profile Applicability, etc.).
    """
    recommendations: List[Dict[str, str]] = []
    lines: List[str] = full_text.splitlines()
    current_recommendation: Dict[str, str] = {}
    current_index: int = 0

    while current_index < len(lines):
        line: str = lines[current_index].strip()
        line = remove_page_numbers(line)

        # Detect a recommendation title line
        title_match = TITLE_PATTERN.match(line)
        if title_match:
            # Check if next lines contain 'Profile Applicability:' => indicates a valid rec
            if find_profile_applicability(lines, current_index):
                # Save previous recommendation if it exists
                if current_recommendation:
                    recommendations.append(current_recommendation)
                current_recommendation = {
                    'Number': title_match.group(1),
                    'Level': title_match.group(2) or '',
                    'Title': title_match.group(3),
                }
                # Capture multi-line titles
                while (current_index + 1 < len(lines)
                       and not any(lines[current_index + 1].strip().startswith(sec) for sec in SECTIONS_WITHOUT_CIS)
                       and not TITLE_PATTERN.match(lines[current_index + 1].strip())):
                    current_index += 1
                    current_recommendation['Title'] += " " + lines[current_index].strip()

        # Extract standard sections (Profile Applicability, Description, etc.)
        for sec in SECTIONS_WITHOUT_CIS:
            if line.startswith(sec):
                content, next_index = extract_section(lines, current_index, sec)
                # e.g. "Additional Information:" -> key = "Additional Information"
                current_recommendation[sec[:-1]] = content
                current_index = next_index - 1
                break

        current_index += 1

    # Add the last recommendation if any
    if current_recommendation:
        recommendations.append(current_recommendation)

    # Remove duplicates based on (Number, Title) in case of accidental repeats
    unique_recommendations = {(rec['Number'], rec['Title']): rec for rec in recommendations}
    return list(unique_recommendations.values())

# -----------------------------------------------------------------------------------
# Output Generation (CSV/Excel/JSON)
# -----------------------------------------------------------------------------------

def write_output(
    recommendations: List[Dict[str, str]],
    output_file: Path,
    output_format: str,
    title: str,
    version: str
) -> None:
    """
    Writes the extracted recommendations to CSV, Excel, or JSON format.
    
    Args:
        recommendations : List of recommendation dicts.
        output_file     : Output file path.
        output_format   : "csv", "excel", or "json".
        title           : Document title (extracted from PDF).
        version         : Document version (extracted from PDF).
    """
    logging.info(f"Writing output to {output_file} in {output_format.upper()} format...")
    
    if output_format == 'csv':
        headers: List[str] = ['Compliance Status', 'Number', 'Level', 'Title']
        headers += [sec[:-1] for sec in SECTIONS_WITHOUT_CIS]  # remove trailing colon
        try:
            with output_file.open(mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file, delimiter='|')
                writer.writerow([title if title else "CIS Benchmark Document"])
                writer.writerow([version if version else ""])
                writer.writerow([])  # Empty line for spacing
                writer.writerow(headers)
                for recommendation in recommendations:
                    recommendation['Compliance Status'] = 'To Review'
                    row = [recommendation.get(header, '') for header in headers]
                    writer.writerow(row)
        except Exception as e:
            logging.error(f"Error writing CSV output: {e}")
            raise

    elif output_format == 'excel':
        headers: List[str] = ['Compliance Status', 'Number', 'Level', 'Title']
        headers += [sec[:-1] for sec in SECTIONS_WITHOUT_CIS]  # remove trailing colon
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Recommendations"
        sheet["A1"] = title if title else "CIS Benchmark Document"
        sheet["A1"].font = Font(size=14, bold=True)
        sheet["A2"] = version if version else ""
        sheet["A2"].font = Font(size=12, italic=True)
        sheet.append([""] * len(headers))
        sheet.append(headers)
        for recommendation in recommendations:
            recommendation['Compliance Status'] = 'To Review'
            row = [recommendation.get(header, '') for header in headers]
            sheet.append(row)
        dv = DataValidation(type="list", formula1='"Compliant,Non-Compliant,To Review"', showDropDown=False)
        sheet.add_data_validation(dv)
        start_row = 5
        end_row = len(recommendations) + start_row
        for row_idx in range(start_row, end_row):
            dv.add(sheet[f"A{row_idx}"])
        compliant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        non_compliant_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        to_review_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        compliant_rule = FormulaRule(formula=[f'$A{start_row}="Compliant"'], fill=compliant_fill)
        non_compliant_rule = FormulaRule(formula=[f'$A{start_row}="Non-Compliant"'], fill=non_compliant_fill)
        to_review_rule = FormulaRule(formula=[f'$A{start_row}="To Review"'], fill=to_review_fill)
        sheet.conditional_formatting.add(f"A{start_row}:A{end_row}", compliant_rule)
        sheet.conditional_formatting.add(f"A{start_row}:A{end_row}", non_compliant_rule)
        sheet.conditional_formatting.add(f"A{start_row}:A{end_row}", to_review_rule)
        last_column = get_column_letter(len(headers))
        table_range = f"A4:{last_column}{end_row - 1}"
        table = Table(displayName="CISRecommendations", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium9",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 50
        for col in range(5, len(headers) + 1):
            col_letter = get_column_letter(col)
            sheet.column_dimensions[col_letter].width = 10
        try:
            workbook.save(str(output_file))
        except Exception as e:
            logging.error(f"Error saving Excel file: {e}")
            raise

    elif output_format == 'json':
        try:
            # Add the new fields to each recommendation
            for recommendation in recommendations:
                recommendation['Implemented'] = ''
                recommendation['Reasoning'] = ''
            # Create a JSON object with document information and recommendations
            data = {
                "document_title": title if title else "CIS Benchmark Document",
                "document_version": version,
                "recommendations": recommendations
            }
            with output_file.open("w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            logging.error(f"Error writing JSON output: {e}")
            raise

    logging.info(f"Finished writing {len(recommendations)} recommendations to {output_file}.")

# -----------------------------------------------------------------------------------
# Main Function
# -----------------------------------------------------------------------------------

def main() -> None:
    """
    Main entry point.
    Parses command-line arguments, extracts recommendations from the PDF,
    and writes the results to CSV, Excel, or JSON.
    """
    parser = argparse.ArgumentParser(description="Extract and format recommendations from a CIS Benchmark PDF.")
    parser.add_argument("-i", "--input", required=True, type=Path, help="Input PDF file.")
    parser.add_argument("-o", "--output", type=Path,
                        help="Output file (default: same as input file name with .csv, .xlsx, or .json).")
    parser.add_argument("-f", "--format", choices=['csv', 'excel', 'json'], default='excel',
                        help="Output format (csv, excel, or json).")
    parser.add_argument("--start_page", type=int, default=10,
                        help="Page number to start extraction (default: 10).")
    parser.add_argument("--log_level", type=str, default="INFO",
                        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
                        help="Logging level (default: INFO).")
    args = parser.parse_args()

    # Configure logging level
    logging.getLogger().setLevel(args.log_level.upper())

    # Prepare file paths
    input_file: Path = args.input
    output_format: str = args.format
    base_name: str = input_file.stem
    extension: str = "csv" if output_format == "csv" else ("xlsx" if output_format == "excel" else "json")
    output_file: Path = args.output if args.output else Path(generate_unique_filename(base_name, extension))

    # Extract title and version from the PDF
    title, version = extract_title_and_version(input_file)

    # Read text from PDF, starting at the user-specified page
    pdf_text = read_pdf(input_file, start_page=args.start_page)

    # Extract recommendations from the raw PDF text
    recommendations = extract_recommendations(pdf_text)

    # Write the extracted data to CSV, Excel, or JSON
    write_output(recommendations, output_file, output_format, title, version)

if __name__ == "__main__":
    main()